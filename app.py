import os
from flask import Flask, render_template, redirect, url_for, flash, request
from io import BytesIO
from flask import send_file

from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta
import random
import pandas as pd        # <-- Ye line add karo
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from models import db, User, Feedback, PasswordReset, UploadedFile, TeacherSubject, AdminSettings
from sqlalchemy import and_, or_




from flask_login import login_required




from models import UploadedFile  # tumhara model jahan files store hai

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.secret_key = "secretkey"



app = Flask(__name__)


app.config['UPLOAD_FOLDER'] = os.path.join(os.getcwd(), 'uploads')
if not os.path.exists(app.config['UPLOAD_FOLDER']):
    os.makedirs(app.config['UPLOAD_FOLDER'])

app.config['SECRET_KEY'] = 'college-feedback-secret-key-2024'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///feedback.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)
login_manager = LoginManager(app)
login_manager.login_view = 'login'

BRANCHES = ['CSE(AI&DS)', 'Civil', 'Electrical']
BRANCH_ALIASES = {
    'cseai&ds': 'CSE(AI&DS)',
    'cseaiandds': 'CSE(AI&DS)',
    'cseaids': 'CSE(AI&DS)',
    'cse-ai&ds': 'CSE(AI&DS)',
    'cse-ai-ds': 'CSE(AI&DS)',
    'civil': 'Civil',
    'electrical': 'Electrical',
}


@app.after_request
def add_no_cache_headers(response):
    if request.path.startswith('/student') or request.path.startswith('/feedback'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

def get_feedback_settings():
    settings = AdminSettings.query.first()
    if not settings:
        settings = AdminSettings(allow_feedback=True)
        db.session.add(settings)
        db.session.commit()
    return settings


def normalize_col(col):
    return str(col).strip().lower().replace(" ", "_").replace("-", "_")


def branch_key(branch):
    return (
        str(branch or '')
        .strip()
        .lower()
        .replace(' ', '')
        .replace('(', '')
        .replace(')', '')
        .replace('-', '')
        .replace('_', '')
        .replace('and', '&')
    )


def normalize_branch(branch):
    text = str(branch or '').strip()
    if not text:
        return ''
    return BRANCH_ALIASES.get(branch_key(text), text)


def branch_key_column(column):
    expr = db.func.lower(db.func.trim(column))
    for old, new in [
        (' ', ''),
        ('(', ''),
        (')', ''),
        ('-', ''),
        ('_', ''),
        ('and', '&'),
    ]:
        expr = db.func.replace(expr, old, new)
    return expr


def cell(row, *names, default=''):
    for name in names:
        key = normalize_col(name)
        if key in row:
            value = str(row.get(key, '')).strip()
            if value.lower() != 'nan':
                return value
    return default


def parse_semester(value, default=1):
    text = str(value).strip().lower()
    digits = ''.join(ch for ch in text if ch.isdigit())
    if digits:
        return int(digits)
    return default


def get_fresh_current_user():
    return User.query.populate_existing().filter_by(id=current_user.id).first()


def upload_path(filename):
    return os.path.join(app.config['UPLOAD_FOLDER'], filename)


def mark_file_status(files):
    for file in files:
        file.physical_exists = os.path.exists(upload_path(file.filename))
    return files

def resolve_unique_email(email, username, user_id=None):
    email = (email or '').strip() or f'{username}@college.com'
    existing = User.query.filter(User.email == email)
    if user_id is not None:
        existing = existing.filter(User.id != user_id)

    if not existing.first():
        return email

    fallback = f'{username}@college.com'
    existing = User.query.filter(User.email == fallback)
    if user_id is not None:
        existing = existing.filter(User.id != user_id)

    if not existing.first():
        return fallback

    return f'{username}_{user_id or "new"}@college.com'


def import_students_from_excel(file_path, uploaded_file_id):
    df = pd.read_excel(file_path, dtype=str).fillna('')
    df.columns = [normalize_col(col) for col in df.columns]

    imported = 0
    updated = 0

    for _, row in df.iterrows():
        username = cell(row, 'username', 'user_name', 'roll_number', 'roll no', 'roll')
        if not username:
            continue

        password = cell(row, 'password', default='student@123') or 'student@123'
        name = cell(row, 'name', 'student_name', default=username)
        email = cell(row, 'email', 'e_mail', 'mail', default=f'{username}@college.com')
        branch = normalize_branch(cell(row, 'branch'))
        semester = parse_semester(cell(row, 'semester', 'sem', default='1'))
        academic_year = cell(row, 'academic_year', 'academic year', 'year', default='2024-2025')
        roll_number = cell(row, 'roll_number', 'roll number', 'roll_no', 'roll no', default=username)

        user = User.query.filter_by(username=username).first()
        email = resolve_unique_email(email, username, user.id if user else None)

        if user:
            if user.imported_file_id is None:
                Feedback.query.filter_by(
                    student_id=user.id,
                    semester=semester
                ).delete(synchronize_session=False)

            user.password = generate_password_hash(password)
            user.name = name
            user.email = email
            user.role = 'student'
            user.branch = branch
            user.semester = semester
            user.academic_year = academic_year
            user.roll_number = roll_number
            user.imported_file_id = uploaded_file_id
            updated += 1
        else:
            user = User(
                username=username,
                password=generate_password_hash(password),
                role='student',
                name=name,
                email=email,
                branch=branch,
                semester=semester,
                academic_year=academic_year,
                roll_number=roll_number,
                imported_file_id=uploaded_file_id
            )
            db.session.add(user)
            imported += 1

    db.session.commit()
    return imported, updated



def import_teachers_from_excel(file_path, uploaded_file_id):
    df = pd.read_excel(file_path, dtype=str).fillna('')
    df.columns = [normalize_col(col) for col in df.columns]

    imported_users = 0
    imported_subjects = 0
    updated_subjects = 0
    prepared_rows = []
    teacher_ids = set()
    branches = set()
    updated_login_teacher_ids = set()

    for _, row in df.iterrows():
        username = cell(row, 'username', 'user name', 'teacher username').strip()
        if not username:
            continue

        password = cell(row, 'password', default='teacher123') or 'teacher123'
        email = cell(row, 'email', 'e-mail', 'mail')
        name = cell(row, 'name', 'teacher name', default=username) or username
        subject = cell(row, 'subject', 'subject name')
        subject_code = cell(row, 'subject_code', 'subject code', 'code')
        branch = normalize_branch(cell(row, 'branch'))
        semester = parse_semester(cell(row, 'semester', 'sem', default='1'))

        normalized_subject_code_branch = normalize_branch(subject_code)
        if branch not in BRANCHES and normalized_subject_code_branch in BRANCHES:
            branch, subject_code = normalized_subject_code_branch, branch

        if not subject or not branch:
            continue

        user = User.query.filter(
            db.func.lower(db.func.trim(User.username)) == username.strip().lower()
        ).first()
        if not user and email:
            user = User.query.filter(
                db.func.lower(db.func.trim(User.email)) == email.strip().lower()
            ).first()
        email = resolve_unique_email(email, username.replace(' ', '').lower(), user.id if user else None)

        if user:
            if user.imported_file_id is None:
                Feedback.query.filter_by(teacher_id=user.id).delete(synchronize_session=False)

            if user.id not in updated_login_teacher_ids:
                user.password = generate_password_hash(password)
                if not User.query.filter(
                    db.func.lower(db.func.trim(User.username)) == username.strip().lower(),
                    User.id != user.id
                ).first():
                    user.username = username
                updated_login_teacher_ids.add(user.id)
            user.role = 'teacher'
            user.name = name
            user.email = email
            user.subject = subject
            user.imported_file_id = uploaded_file_id
            db.session.flush()
        else:
            user = User(
                username=username,
                password=generate_password_hash(password),
                role='teacher',
                name=name,
                email=email,
                subject=subject,
                imported_file_id=uploaded_file_id
            )
            db.session.add(user)
            db.session.flush()
            updated_login_teacher_ids.add(user.id)
            imported_users += 1

        teacher_ids.add(user.id)
        branches.add(branch)
        prepared_rows.append({
            'teacher_id': user.id,
            'subject': subject,
            'subject_code': subject_code,
            'branch': branch,
            'semester': semester
        })

    if teacher_ids and branches:
        TeacherSubject.query.filter(
            TeacherSubject.teacher_id.in_(teacher_ids)
        ).delete(synchronize_session=False)

    seen = set()
    for item in prepared_rows:
        key = (
            item['teacher_id'],
            item['branch'].strip().lower(),
            item['semester'],
            item['subject'].strip().lower()
        )
        if key in seen:
            updated_subjects += 1
            continue
        seen.add(key)

        teacher_subject = TeacherSubject(
            teacher_id=item['teacher_id'],
            subject=item['subject'],
            subject_code=item['subject_code'],
            branch=item['branch'],
            semester=item['semester']
        )
        db.session.add(teacher_subject)
        imported_subjects += 1

    db.session.commit()
    return imported_users, imported_subjects, updated_subjects

def get_current_teacher_semester(teacher_id):
    semesters = [
        sem for (sem,) in db.session.query(TeacherSubject.semester)
        .filter(TeacherSubject.teacher_id == teacher_id, TeacherSubject.semester.isnot(None))
        .distinct()
        .all()
        if sem is not None
    ]
    return max(semesters) if semesters else None


def attach_subject_codes(feedbacks):
    for f in feedbacks:
        f.average_rating = round(f.total / 12, 2)
        teacher_subject = TeacherSubject.query.filter(
            TeacherSubject.teacher_id == f.teacher_id,
            TeacherSubject.semester == f.semester,
            db.func.lower(db.func.trim(TeacherSubject.subject)) == (f.subject or '').strip().lower()
        ).first()
        f.subject_code = teacher_subject.subject_code if teacher_subject else find_subject_code(f.teacher_id, f.subject)
    return feedbacks


def feedback_rows(feedbacks):
    rows = []
    for f in feedbacks:
        rows.append({
            'Student': f.student.name if f.student else '',
            'Roll No': f.student.roll_number if f.student else '',
            'Branch': f.student.branch if f.student else '',
            'Semester': f.semester,
            'Teacher': f.teacher.name if f.teacher else '',
            'Subject Code': getattr(f, 'subject_code', find_subject_code(f.teacher_id, f.subject)),
            'Subject': f.subject,
            'Q1': f.q1,
            'Q2': f.q2,
            'Q3': f.q3,
            'Q4': f.q4,
            'Q5': f.q5,
            'Q6': f.q6,
            'Q7': f.q7,
            'Q8': f.q8,
            'Q9': f.q9,
            'Q10': f.q10,
            'Q11': f.q11,
            'Q12': f.q12,
            'Total': f.total,
            'Average': round(f.total / 12, 2),
            'Date': f.created_at.strftime('%d-%m-%Y %I:%M %p') if f.created_at else ''
        })
    return rows


def export_feedbacks_response(feedbacks, download_name):
    output = BytesIO()
    pd.DataFrame(feedback_rows(attach_subject_codes(feedbacks))).to_excel(output, index=False)
    output.seek(0)
    return send_file(output, as_attachment=True, download_name=download_name)


def rating_label(value):
    if value >= 4.5:
        return 'Excellent'
    if value >= 3.5:
        return 'Very Good'
    if value >= 2.5:
        return 'Good'
    if value >= 1.5:
        return 'Poor'
    return 'Very Poor'


def feedback_average_summary(feedbacks):
    teacher_rows = {}
    teacher_semester_rows = {}
    subject_rows = {}

    for fb in feedbacks:
        rating = (fb.total or 0) / 12
        teacher_name = fb.teacher.name or fb.teacher.username if fb.teacher else 'Unknown Teacher'
        teacher_key = fb.teacher_id

        teacher_row = teacher_rows.setdefault(teacher_key, {
            'teacher': teacher_name,
            'feedback_count': 0,
            'rating_total': 0,
            'subjects': set(),
            'branches': set(),
            'semesters': set()
        })
        teacher_row['feedback_count'] += 1
        teacher_row['rating_total'] += rating
        teacher_row['subjects'].add(fb.subject or '-')
        teacher_row['branches'].add(fb.student.branch if fb.student else '-')
        teacher_row['semesters'].add(fb.semester)

        branch = fb.student.branch if fb.student else '-'
        subject = fb.subject or '-'
        subject_code = getattr(fb, 'subject_code', None) or find_subject_code(fb.teacher_id, fb.subject)
        teacher_semester_key = (
            teacher_key,
            teacher_name.strip().lower(),
            (branch or '-').strip().lower(),
            fb.semester
        )
        teacher_semester_row = teacher_semester_rows.setdefault(teacher_semester_key, {
            'teacher': teacher_name,
            'branch': branch or '-',
            'semester': fb.semester,
            'feedback_count': 0,
            'rating_total': 0,
            'subjects': set()
        })
        teacher_semester_row['feedback_count'] += 1
        teacher_semester_row['rating_total'] += rating
        teacher_semester_row['subjects'].add(subject)

        subject_key = (
            teacher_key,
            teacher_name.strip().lower(),
            subject.strip().lower(),
            (subject_code or '-').strip().lower(),
            (branch or '-').strip().lower(),
            fb.semester
        )

        subject_row = subject_rows.setdefault(subject_key, {
            'teacher': teacher_name,
            'subject': subject,
            'subject_code': subject_code or '-',
            'branch': branch or '-',
            'semester': fb.semester,
            'feedback_count': 0,
            'rating_total': 0
        })
        subject_row['feedback_count'] += 1
        subject_row['rating_total'] += rating

    for row in teacher_rows.values():
        row['average_rating'] = round(row['rating_total'] / row['feedback_count'], 2) if row['feedback_count'] else 0
        row['rating_label'] = rating_label(row['average_rating'])
        row['subjects'] = ', '.join(sorted(row['subjects']))
        row['branches'] = ', '.join(sorted(row['branches']))
        row['semesters'] = ', '.join(str(sem) for sem in sorted(row['semesters']))

    for row in teacher_semester_rows.values():
        row['average_rating'] = round(row['rating_total'] / row['feedback_count'], 2) if row['feedback_count'] else 0
        row['rating_label'] = rating_label(row['average_rating'])
        row['subjects'] = ', '.join(sorted(row['subjects']))

    for row in subject_rows.values():
        row['average_rating'] = round(row['rating_total'] / row['feedback_count'], 2) if row['feedback_count'] else 0
        row['rating_label'] = rating_label(row['average_rating'])

    overall_average = round(
        sum((fb.total or 0) / 12 for fb in feedbacks) / len(feedbacks),
        2
    ) if feedbacks else 0
    return {
        'overall_average': overall_average,
        'overall_label': rating_label(overall_average) if feedbacks else '-',
        'teacher_rows': sorted(teacher_rows.values(), key=lambda item: item['teacher'].lower()),
        'teacher_semester_rows': sorted(
            teacher_semester_rows.values(),
            key=lambda item: (
                item['branch'].lower(),
                item['semester'] or 0,
                item['teacher'].lower()
            )
        ),
        'subject_rows': sorted(
            subject_rows.values(),
            key=lambda item: (
                item['branch'].lower(),
                item['semester'] or 0,
                item['teacher'].lower(),
                item['subject'].lower()
            )
        )
    }

def find_subject_code(teacher_id, subject):
    feedback_subject = (subject or '').strip().lower()

    teacher_subjects = TeacherSubject.query.filter_by(
        teacher_id=teacher_id
    ).all()

    for ts in teacher_subjects:
        ts_subject = (ts.subject or '').strip().lower()
        if ts_subject == feedback_subject:
            return ts.subject_code or '-'

    return '-'




@login_manager.user_loader
def load_user(user_id):
    return db.session.get(User, int(user_id))

from flask import send_from_directory

@app.route('/open_file/<int:file_id>')
@login_required
def open_file(file_id):
    file = db.session.get(UploadedFile, file_id)
    if not file:
        flash("File not found!", "danger")
        return redirect(url_for('superadmin_dashboard'))

    upload_folder = os.path.join(app.root_path, 'uploads')
    return send_from_directory(upload_folder, file.filename, as_attachment=True)

# ========================
# HOME ROUTE
# ========================
@app.route('/')
def home():
    if current_user.is_authenticated:
        role = (getattr(current_user, 'role', None) or '').strip().lower()
        sub_role = (getattr(current_user, 'sub_role', None) or '').strip().lower()

        if role == 'admin':
            if sub_role == 'superadmin':
                return redirect(url_for('superadmin_dashboard'))
            return redirect(url_for('admin_dashboard'))

        if role == 'teacher':
            return redirect(url_for('teacher_dashboard'))

        if role == 'student':
            return redirect(url_for('student_dashboard'))

        logout_user()
        return redirect(url_for('login'))

    return redirect(url_for('login'))


# ========================
# LOGIN / LOGOUT
# ========================
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        print("LOGIN FORM:", request.form)

        username = (request.form.get('username') or '').strip()
        password = request.form.get('password')
        selected_role = (request.form.get('role') or '').strip().lower()

        user = User.query.filter_by(username=username).first()
        print("USER FOUND:", user)
        print("SELECTED ROLE:", selected_role)


        if user and check_password_hash(user.password, password):
            user_role = (user.role or '').strip().lower()

            if user_role == selected_role:
                login_user(user)

                if user_role == 'student':
                    return redirect(url_for('student_dashboard'))
                elif user_role == 'teacher':
                    return redirect(url_for('teacher_dashboard'))
                elif user_role == 'admin':
                    return redirect(url_for('home'))

            else:
                flash("You selected wrong login type!", "danger")
        else:
            flash("Invalid username or password!", "danger")

    return render_template('login.html')




@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Logged out successfully!', 'info')
    return redirect(url_for('login'))

# ========================
# FORGOT / RESET PASSWORD
# ========================
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        user = User.query.filter_by(username=request.form.get('username')).first()
        if user:
            code = ''.join(random.choices('0123456789', k=6))
            reset = PasswordReset(
                user_id=user.id,
                reset_code=code,
                expires_at=datetime.utcnow() + timedelta(minutes=30)
            )
            db.session.add(reset)
            db.session.commit()
            flash(f'Reset code: {code}', 'info')
            return redirect(url_for('reset_password'))
        flash('Username not found!', 'danger')
    return render_template('forgot_password.html')

@app.route('/reset_password', methods=['GET', 'POST'])
def reset_password():
    if request.method == 'POST':
        reset = PasswordReset.query.filter_by(
            reset_code=request.form.get('reset_code'),
            is_used=False
        ).first()
        if reset and reset.expires_at > datetime.utcnow():
            user = db.session.get(User, reset.user_id)
            user.password = generate_password_hash(request.form.get('new_password'))
            reset.is_used = True
            db.session.commit()
            flash('Password reset successful! Login now.', 'success')
            return redirect(url_for('login'))
        flash('Invalid or expired reset code!', 'danger')
    return render_template('reset_password.html')

# ========================
# STUDENT DASHBOARD
# ========================
@app.route('/student')
@login_required
def student_dashboard():
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    if role != 'student':
        return redirect(url_for('home'))

    student = get_fresh_current_user()
    if not student:
        logout_user()
        return redirect(url_for('login'))

    student_branch_key = branch_key(normalize_branch(student.branch))
    teacher_subjects = TeacherSubject.query.join(
        User, TeacherSubject.teacher_id == User.id
    ).filter(
        User.role == 'teacher',
        branch_key_column(TeacherSubject.branch) == student_branch_key,
        TeacherSubject.semester == student.semester
    ).all()

    submitted_feedbacks = Feedback.query.filter_by(
        student_id=student.id,
        semester=student.semester
    ).all()

    submitted_keys = [(f.teacher_id, f.subject) for f in submitted_feedbacks]

    feedback_settings = get_feedback_settings()
    now = datetime.now()

    feedback_open = (
        feedback_settings.feedback_start
        and feedback_settings.feedback_end
        and feedback_settings.feedback_start <= now <= feedback_settings.feedback_end
    )

    return render_template(
        'student/dashboard.html',
        student=student,
        teacher_subjects=teacher_subjects,
        submitted_keys=submitted_keys,
        feedback_settings=feedback_settings,
        feedback_open=feedback_open
    )




@app.route('/submit_feedback/<int:teacher_id>', methods=['POST'])
@login_required
def submit_feedback(teacher_id):
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    if role != 'student':
        return redirect(url_for('home'))

    student = get_fresh_current_user()
    if not student:
        logout_user()
        return redirect(url_for('login'))

    subject = request.form.get('subject')
    settings = get_feedback_settings()
    now = datetime.now()

    if not settings.feedback_start or not settings.feedback_end or not (settings.feedback_start <= now <= settings.feedback_end):
       flash('Feedback is currently closed.', 'danger')
       return redirect(url_for('student_dashboard'))

    assigned_subject = TeacherSubject.query.filter(
        TeacherSubject.teacher_id == teacher_id,
        branch_key_column(TeacherSubject.branch) == branch_key(normalize_branch(student.branch)),
        TeacherSubject.semester == student.semester,
        db.func.lower(db.func.trim(TeacherSubject.subject)) == (subject or '').strip().lower()
    ).first()

    if not assigned_subject:
        flash('This teacher is not assigned for your current semester.', 'danger')
        return redirect(url_for('student_dashboard'))

    existing = Feedback.query.filter_by(
        student_id=student.id,
        teacher_id=teacher_id,
        subject=subject,
        semester=student.semester
    ).first()

    if existing:
        flash('Feedback already submitted for this teacher this semester!', 'warning')
        return redirect(url_for('student_dashboard'))

    answers = {}
    total = 0

    for i in range(1, 13):
        value = request.form.get(f'q{i}')
        if not value:
            flash('Please answer all questions!', 'danger')
            return redirect(url_for('feedback_form', teacher_id=teacher_id, subject=subject))

        value = int(value)
        answers[f'q{i}'] = value
        total += value

    new_feedback = Feedback(
        student_id=student.id,
        teacher_id=teacher_id,
        subject=subject,
        q1=answers['q1'],
        q2=answers['q2'],
        q3=answers['q3'],
        q4=answers['q4'],
        q5=answers['q5'],
        q6=answers['q6'],
        q7=answers['q7'],
        q8=answers['q8'],
        q9=answers['q9'],
        q10=answers['q10'],
        q11=answers['q11'],
        q12=answers['q12'],
        total=total,
        semester=student.semester,
        academic_year=student.academic_year or '2024-2025'
    )

    db.session.add(new_feedback)
    db.session.commit()

    flash('Feedback submitted successfully!', 'success')
    return redirect(url_for('student_dashboard'))

@app.route('/feedback/<int:teacher_id>')
@login_required
def feedback_form(teacher_id):
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    if role != 'student':
        return redirect(url_for('home'))

    student = get_fresh_current_user()
    if not student:
        logout_user()
        return redirect(url_for('login'))

    teacher = User.query.get_or_404(teacher_id)
    subject = request.args.get('subject')

    assigned_subject = TeacherSubject.query.filter(
        TeacherSubject.teacher_id == teacher_id,
        branch_key_column(TeacherSubject.branch) == branch_key(normalize_branch(student.branch)),
        TeacherSubject.semester == student.semester,
        db.func.lower(db.func.trim(TeacherSubject.subject)) == (subject or '').strip().lower()
    ).first()

    if not assigned_subject:
        flash('This teacher is not assigned for your current semester.', 'danger')
        return redirect(url_for('student_dashboard'))

    existing = Feedback.query.filter_by(
        student_id=student.id,
        teacher_id=teacher_id,
        subject=subject,
        semester=student.semester
    ).first()

    if existing:
        flash('You already submitted feedback for this teacher.', 'warning')
        return redirect(url_for('student_dashboard'))

    return render_template('student/feedback.html', teacher=teacher, subject=subject)

# ========================
# TEACHER DASHBOARD
# ========================
@app.route('/teacher')
@login_required
def teacher_dashboard():
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    if role != 'teacher':
        return redirect(url_for('home'))

    selected_branch = request.args.get('branch', '')
    selected_semester = request.args.get('semester', '')
    selected_subject = request.args.get('subject', '')

    base_query = Feedback.query.join(User, Feedback.student_id == User.id).filter(
        Feedback.teacher_id == current_user.id,
        Feedback.semester == User.semester
    )

    semester_rows = base_query.with_entities(Feedback.semester).distinct().order_by(Feedback.semester).all()
    current_semesters = [row[0] for row in semester_rows]

    query = base_query

    if selected_branch:
        query = query.filter(User.branch == selected_branch)

    if selected_semester:
        query = query.filter(Feedback.semester == int(selected_semester))

    if selected_subject:
        query = query.filter(Feedback.subject == selected_subject)

    feedbacks = attach_subject_codes(query.order_by(Feedback.created_at.desc()).all())

    total = len(feedbacks)
    avg = sum((f.total / 12) for f in feedbacks) / total if total else 0

    excellent = sum(1 for f in feedbacks if (f.total / 12) >= 4.5)
    very_good = sum(1 for f in feedbacks if 3.5 <= (f.total / 12) < 4.5)
    good = sum(1 for f in feedbacks if 2.5 <= (f.total / 12) < 3.5)
    poor = sum(1 for f in feedbacks if 1.5 <= (f.total / 12) < 2.5)
    very_poor = sum(1 for f in feedbacks if (f.total / 12) < 1.5)

    teacher_subjects = TeacherSubject.query.filter_by(teacher_id=current_user.id).all()

    return render_template(
        'teacher/teacher_dashboard.html',
        feedbacks=feedbacks,
        total_feedbacks=total,
        average_rating=round(avg, 1),
        average_rating_label=rating_label(avg) if total else '-',
        excellent_count=excellent,
        very_good_count=very_good,
        good_count=good,
        poor_count=poor,
        very_poor_count=very_poor,
        branches=BRANCHES,
        teacher_subjects=teacher_subjects,
        selected_subject=selected_subject,
        selected_semester=selected_semester,
        current_semesters=current_semesters
    )


@app.route('/teacher/old')
@login_required
def teacher_old_feedbacks():
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    if role != 'teacher':
        return redirect(url_for('home'))

    selected_branch = request.args.get('branch', '')
    selected_semester = request.args.get('semester', '')
    selected_subject = request.args.get('subject', '')

    base_query = Feedback.query.join(User, Feedback.student_id == User.id).filter(
        Feedback.teacher_id == current_user.id,
        Feedback.semester != User.semester
    )

    semester_rows = base_query.with_entities(Feedback.semester).distinct().order_by(Feedback.semester).all()
    old_semesters = [row[0] for row in semester_rows]

    query = base_query

    if selected_branch:
        query = query.filter(User.branch == selected_branch)

    if selected_semester:
        query = query.filter(Feedback.semester == int(selected_semester))

    if selected_subject:
        query = query.filter(Feedback.subject == selected_subject)

    feedbacks = attach_subject_codes(query.order_by(Feedback.created_at.desc()).all())
    teacher_subjects = TeacherSubject.query.filter_by(teacher_id=current_user.id).all()

    return render_template(
        'teacher/old_feedbacks.html',
        feedbacks=feedbacks,
        branches=BRANCHES,
        teacher_subjects=teacher_subjects,
        old_semesters=old_semesters,
        selected_subject=selected_subject,
        selected_semester=selected_semester
    )


@app.route('/teacher/old/export')
@login_required
def export_teacher_old_feedbacks():
    if current_user.role != 'teacher':
        return redirect(url_for('home'))

    query = Feedback.query.join(User, Feedback.student_id == User.id).filter(
        Feedback.teacher_id == current_user.id,
        Feedback.semester != User.semester
    )

    return export_feedbacks_response(
        query.order_by(Feedback.created_at.desc()).all(),
        f'old_teacher_feedback_{current_user.id}.xlsx'
    )


@app.route('/teacher/old/delete', methods=['POST'])
@login_required
def delete_teacher_old_feedbacks():
    if current_user.role != 'teacher':
        return redirect(url_for('home'))

    old_ids = [row[0] for row in db.session.query(Feedback.id)
        .join(User, Feedback.student_id == User.id)
        .filter(Feedback.teacher_id == current_user.id, Feedback.semester != User.semester)
        .all()]

    deleted = 0
    if old_ids:
        deleted = Feedback.query.filter(Feedback.id.in_(old_ids)).delete(synchronize_session=False)
        db.session.commit()

    flash(f'Old feedback deleted: {deleted}', 'success')
    return redirect(url_for('teacher_old_feedbacks'))# ========================
# BRANCH ADMIN DASHBOARD
# ========================
@app.route('/admin')
@login_required
def admin_dashboard():
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'branchadmin':
        return redirect(url_for('home'))
    selected_branch = current_user.branch
    selected_semester = request.args.get('semester', None, type=int)
    selected_teacher = request.args.get('teacher', None, type=int)

    students_query = User.query.filter_by(role='student', branch=selected_branch)
    if selected_semester:
        students_query = students_query.filter_by(semester=selected_semester)
    students = students_query.all()
    teachers = User.query.filter_by(role='teacher', branch=selected_branch).all()

    feedback_query = Feedback.query.join(User, Feedback.student_id==User.id).filter(User.branch==selected_branch)
    if selected_semester:
        feedback_query = feedback_query.filter(Feedback.semester==selected_semester)
    if selected_teacher:
        feedback_query = feedback_query.filter(Feedback.teacher_id==selected_teacher)
    feedbacks = feedback_query.all()

    total = len(feedbacks)
    avg = sum(f.rating for f in feedbacks)/total if total else 0
    excellent = sum(1 for f in feedbacks if f.rating==5)
    good = sum(1 for f in feedbacks if f.rating==4)
    average = sum(1 for f in feedbacks if f.rating==3)
    poor = sum(1 for f in feedbacks if f.rating==2)
    very_poor = sum(1 for f in feedbacks if f.rating==1)
    feedback_settings = get_feedback_settings()


    return render_template('admin/dashboard.html', students=students, teachers=teachers, branches=BRANCHES,
                           selected_branch=selected_branch, selected_semester=selected_semester,
                           selected_teacher=selected_teacher, total_feedbacks=total,
                           average_rating=round(avg,1), excellent_count=excellent, good_count=good,
                           average_count=average, poor_count=poor, very_poor_count=very_poor,feedback_settings=feedback_settings)


def get_superadmin_feedback_filters():
    return {
        'branch': request.args.get('branch', ''),
        'semester': request.args.get('semester', ''),
        'subject_code': request.args.get('subject_code', ''),
        'subject': request.args.get('subject', ''),
        'teacher': request.args.get('teacher', '')
    }


def build_superadmin_feedback_query(filters, previous=False):
    feedback_query = Feedback.query.join(User, Feedback.student_id == User.id)

    if previous:
        feedback_query = feedback_query.filter(Feedback.semester != User.semester)
    else:
        feedback_query = feedback_query.filter(Feedback.semester == User.semester)

    if filters['branch']:
        feedback_query = feedback_query.filter(
            db.func.lower(db.func.trim(User.branch)) == filters['branch'].strip().lower()
        )

    if filters['semester']:
        feedback_query = feedback_query.filter(Feedback.semester == int(filters['semester']))

    if filters['teacher']:
        feedback_query = feedback_query.filter(Feedback.teacher_id == int(filters['teacher']))

    if filters['subject_code'] or filters['subject']:
        mapping_query = TeacherSubject.query

        if filters['subject_code']:
            mapping_query = mapping_query.filter(
                db.func.lower(db.func.trim(TeacherSubject.subject_code)) == filters['subject_code'].strip().lower()
            )

        if filters['subject']:
            mapping_query = mapping_query.filter(
                db.func.lower(db.func.trim(TeacherSubject.subject)) == filters['subject'].strip().lower()
            )

        mappings = mapping_query.all()

        if mappings:
            conditions = [
                and_(
                    Feedback.teacher_id == m.teacher_id,
                    db.func.lower(db.func.trim(Feedback.subject)) == (m.subject or '').strip().lower()
                )
                for m in mappings
            ]
            feedback_query = feedback_query.filter(or_(*conditions))
        else:
            feedback_query = feedback_query.filter(Feedback.id == -1)

    return feedback_query


def add_subject_codes_to_feedbacks(feedbacks):
    for fb in feedbacks:
        fb.subject_code = find_subject_code(fb.teacher_id, fb.subject)
    return feedbacks


def student_file_semesters(file_path):
    if not os.path.exists(file_path):
        return set()

    df = pd.read_excel(file_path, dtype=str).fillna('')
    df.columns = [normalize_col(col) for col in df.columns]
    semesters = set()

    for _, row in df.iterrows():
        semester = parse_semester(cell(row, 'semester', 'sem', default=''))
        if semester:
            semesters.add(semester)

    return semesters


def average_summary_rows(summary):
    teacher_rows = []
    for row in summary['teacher_rows']:
        teacher_rows.append({
            'Type': 'Teacher Combined',
            'Branch': row['branches'],
            'Semester': row['semesters'],
            'Teacher': row['teacher'],
            'Subject Code': '',
            'Subject': row['subjects'],
            'Feedbacks': row['feedback_count'],
            'Average Rating': row['average_rating'],
            'Rating': row['rating_label']
        })

    teacher_semester_rows = []
    for row in summary['teacher_semester_rows']:
        teacher_semester_rows.append({
            'Type': 'Teacher Branch Semester Combined',
            'Branch': row['branch'],
            'Semester': row['semester'],
            'Teacher': row['teacher'],
            'Subject Code': '',
            'Subject': row['subjects'],
            'Feedbacks': row['feedback_count'],
            'Average Rating': row['average_rating'],
            'Rating': row['rating_label']
        })

    subject_rows = []
    for row in summary['subject_rows']:
        subject_rows.append({
            'Type': 'Subject Wise',
            'Branch': row['branch'],
            'Semester': row['semester'],
            'Teacher': row['teacher'],
            'Subject Code': row['subject_code'],
            'Subject': row['subject'],
            'Feedbacks': row['feedback_count'],
            'Average Rating': row['average_rating'],
            'Rating': row['rating_label']
        })

    return teacher_rows + teacher_semester_rows + subject_rows


def format_average(value):
    return f"{float(value or 0):.2f}/5"


def write_average_section(ws, start_row, title, headers, rows):
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.font = Font(bold=True, size=12)
    start_row += 1

    header_fill = PatternFill('solid', fgColor='F2F4F7')
    border = Border(
        left=Side(style='thin', color='D9DEE7'),
        right=Side(style='thin', color='D9DEE7'),
        top=Side(style='thin', color='D9DEE7'),
        bottom=Side(style='thin', color='D9DEE7')
    )

    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col_index, value=header)
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(vertical='center')

    for row_offset, row_data in enumerate(rows, start=1):
        for col_index, value in enumerate(row_data, start=1):
            cell = ws.cell(row=start_row + row_offset, column=col_index, value=value)
            cell.border = border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    return start_row + len(rows) + 2


def autosize_sheet_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        column_letter = column_cells[0].column_letter
        for cell in column_cells:
            max_length = max(max_length, len(str(cell.value or '')))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 45)
    for column_letter in ['E', 'F', 'G', 'H']:
        if ws.column_dimensions[column_letter].width < 14:
            ws.column_dimensions[column_letter].width = 14


def export_average_summary_response(summary, download_name, previous=False):
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Average Ratings'

    current_row = 1
    current_row = write_average_section(
        ws,
        current_row,
        'Previous Teacher Branch Semester Combined Average' if previous else 'Teacher Branch Semester Combined Average',
        ['Branch', 'Semester', 'Teacher', 'Subjects', 'Feedbacks', 'Average', 'Rating'],
        [
            [
                row['branch'],
                row['semester'],
                row['teacher'],
                row['subjects'],
                row['feedback_count'],
                format_average(row['average_rating']),
                row['rating_label']
            ]
            for row in summary['teacher_semester_rows']
        ]
    )

    if not previous:
        current_row = write_average_section(
            ws,
            current_row,
            'Teacher Combined Average',
            ['Teacher', 'Branches', 'Semesters', 'Subjects', 'Feedbacks', 'Average', 'Rating'],
            [
                [
                    row['teacher'],
                    row['branches'],
                    row['semesters'],
                    row['subjects'],
                    row['feedback_count'],
                    format_average(row['average_rating']),
                    row['rating_label']
                ]
                for row in summary['teacher_rows']
            ]
        )

    write_average_section(
        ws,
        current_row,
        'Previous Subject Wise Average' if previous else 'Subject Wise Average',
        ['Branch', 'Semester', 'Teacher', 'Subject Code', 'Subject', 'Feedbacks', 'Average', 'Rating'],
        [
            [
                row['branch'],
                row['semester'],
                row['teacher'],
                row['subject_code'],
                row['subject'],
                row['feedback_count'],
                format_average(row['average_rating']),
                row['rating_label']
            ]
            for row in summary['subject_rows']
        ]
    )

    autosize_sheet_columns(ws)
    wb.save(output)
    output.seek(0)
    response = send_file(
        output,
        as_attachment=True,
        download_name=download_name,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response

# ========================
# SUPERADMIN DASHBOARD
# ========================
@app.route('/superadmin_dashboard')
@login_required
def superadmin_dashboard():
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    sub_role = (getattr(current_user, 'sub_role', None) or '').strip().lower()

    if role != 'admin' or sub_role != 'superadmin':
        flash("Access denied!", "danger")
        return redirect(url_for('home'))

    students_files = mark_file_status(UploadedFile.query.filter_by(file_type='student').all())
    teachers_files = mark_file_status(UploadedFile.query.filter_by(file_type='teacher').all())
    admins_files = mark_file_status(UploadedFile.query.filter_by(file_type='admin').all())

    filters = get_superadmin_feedback_filters()
    selected_branch = filters['branch']
    selected_semester = filters['semester']
    selected_subject_code = filters['subject_code']
    selected_subject = filters['subject']
    selected_teacher = filters['teacher']

    feedbacks = add_subject_codes_to_feedbacks(
        build_superadmin_feedback_query(filters).order_by(Feedback.created_at.desc()).all()
    )
    previous_feedbacks = add_subject_codes_to_feedbacks(
        build_superadmin_feedback_query(filters, previous=True).order_by(Feedback.created_at.desc()).all()
    )

    average_summary = feedback_average_summary(feedbacks)
    previous_average_summary = feedback_average_summary(previous_feedbacks)

    branch_options = BRANCHES


    subject_code_options = [
        s[0] for s in db.session.query(TeacherSubject.subject_code)
        .filter(TeacherSubject.subject_code.isnot(None), TeacherSubject.subject_code != '')
        .distinct()
        .all()
    ]

    subject_options = [
        s[0] for s in db.session.query(TeacherSubject.subject)
        .filter(TeacherSubject.subject.isnot(None), TeacherSubject.subject != '')
        .distinct()
        .all()
    ]

    teacher_options = User.query.filter_by(role='teacher').order_by(User.username).all()
    feedback_settings = get_feedback_settings()
    total_students_uploaded = User.query.filter_by(role='student').count()

    return render_template(
        'admin/superadmin_dashboard.html',
        students_files=students_files,
        teachers_files=teachers_files,
        admins_files=admins_files,
        feedback_settings=feedback_settings,
        feedbacks=feedbacks,
        branch_options=branch_options,
        subject_code_options=subject_code_options,
        subject_options=subject_options,
        teacher_options=teacher_options,
        selected_branch=selected_branch,
        selected_semester=selected_semester,
        selected_subject_code=selected_subject_code,
        selected_subject=selected_subject,
        selected_teacher=selected_teacher,
        average_summary=average_summary,
        previous_average_summary=previous_average_summary,
        previous_feedbacks=previous_feedbacks,
        total_students_uploaded=total_students_uploaded
    )

@app.route('/export_superadmin_feedback')
@login_required
def export_superadmin_feedback():
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    sub_role = (getattr(current_user, 'sub_role', None) or '').strip().lower()

    if role != 'admin' or sub_role != 'superadmin':
        flash("Access denied!", "danger")
        return redirect(url_for('home'))

    feedbacks = build_superadmin_feedback_query(
        get_superadmin_feedback_filters()
    ).order_by(Feedback.created_at.desc()).all()

    data = []
    for fb in feedbacks:
        data.append({
            "Student": fb.student.name if fb.student else '',
            "Roll No": fb.student.roll_number if fb.student else '',
            "Branch": fb.student.branch if fb.student else '',
            "Semester": fb.semester,
            "Teacher Username": fb.teacher.username if fb.teacher else '',
            "Teacher Name": fb.teacher.name if fb.teacher else '',
            "Subject Code": find_subject_code(fb.teacher_id, fb.subject),
            "Subject": fb.subject,
            "Total": fb.total,
            "Average": round(fb.total / 12, 2),
            "Date": fb.created_at.strftime('%d-%m-%Y %I:%M %p')
        })

    df = pd.DataFrame(data)
    output = BytesIO()
    df.to_excel(output, index=False)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name='filtered_feedback.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/export_superadmin_average')
@login_required
def export_superadmin_average():
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    sub_role = (getattr(current_user, 'sub_role', None) or '').strip().lower()

    if role != 'admin' or sub_role != 'superadmin':
        flash("Access denied!", "danger")
        return redirect(url_for('home'))

    feedbacks = add_subject_codes_to_feedbacks(
        build_superadmin_feedback_query(get_superadmin_feedback_filters()).all()
    )
    return export_average_summary_response(
        feedback_average_summary(feedbacks),
        'current_average_rating.xlsx'
    )


@app.route('/export_superadmin_previous_average')
@login_required
def export_superadmin_previous_average():
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    sub_role = (getattr(current_user, 'sub_role', None) or '').strip().lower()

    if role != 'admin' or sub_role != 'superadmin':
        flash("Access denied!", "danger")
        return redirect(url_for('home'))

    feedbacks = add_subject_codes_to_feedbacks(
        build_superadmin_feedback_query(get_superadmin_feedback_filters(), previous=True).all()
    )
    return export_average_summary_response(
        feedback_average_summary(feedbacks),
        'previous_average_rating.xlsx',
        previous=True
    )


@app.route('/superadmin_previous_average')
@login_required
def superadmin_previous_average():
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    sub_role = (getattr(current_user, 'sub_role', None) or '').strip().lower()

    if role != 'admin' or sub_role != 'superadmin':
        flash("Access denied!", "danger")
        return redirect(url_for('home'))

    filters = get_superadmin_feedback_filters()
    previous_feedbacks = add_subject_codes_to_feedbacks(
        build_superadmin_feedback_query(filters, previous=True).order_by(Feedback.created_at.desc()).all()
    )
    total_students_uploaded = User.query.filter_by(role='student').count()

    return render_template(
        'admin/previous_average.html',
        previous_feedbacks=previous_feedbacks,
        previous_average_summary=feedback_average_summary(previous_feedbacks),
        selected_branch=filters['branch'],
        selected_semester=filters['semester'],
        selected_subject_code=filters['subject_code'],
        selected_subject=filters['subject'],
        selected_teacher=filters['teacher'],
        total_students_uploaded=total_students_uploaded
    )


@app.route('/delete_superadmin_previous_feedbacks', methods=['POST'])
@login_required
def delete_superadmin_previous_feedbacks():
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    sub_role = (getattr(current_user, 'sub_role', None) or '').strip().lower()

    if role != 'admin' or sub_role != 'superadmin':
        flash("Access denied!", "danger")
        return redirect(url_for('home'))

    filters = get_superadmin_feedback_filters()
    previous_ids = [
        row[0] for row in build_superadmin_feedback_query(
            filters,
            previous=True
        ).with_entities(Feedback.id).all()
    ]

    deleted = 0
    if previous_ids:
        deleted = Feedback.query.filter(Feedback.id.in_(previous_ids)).delete(synchronize_session=False)
        db.session.commit()

    flash(f'Previous feedback deleted: {deleted}', 'success')
    return redirect(url_for(
        'superadmin_previous_average',
        branch=filters['branch'],
        semester=filters['semester'],
        subject_code=filters['subject_code'],
        subject=filters['subject'],
        teacher=filters['teacher']
    ))


@app.route('/add_admin', methods=['POST'])
@login_required
def add_admin():
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    username = request.form['username']
    password = request.form['password']
    sub_role = request.form['sub_role']
    branch = request.form.get('branch') if sub_role=='branchadmin' else None
    hashed_pw = generate_password_hash(password)
    new_admin = User(username=username, password=hashed_pw, role='admin', sub_role=sub_role, branch=branch)
    db.session.add(new_admin)
    db.session.commit()
    flash('New admin added successfully!', 'success')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/delete_admin/<int:admin_id>')
@login_required
def delete_admin(admin_id):
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    admin = db.session.get(User, admin_id)
    if admin and admin.sub_role != 'superadmin':
        db.session.delete(admin)
        db.session.commit()
        flash('Admin deleted successfully!', 'success')
    else:
        flash('Cannot delete superadmin!', 'danger')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/change_admin_password/<int:admin_id>', methods=['POST'])
@login_required
def change_admin_password(admin_id):
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    new_password = request.form['new_password']
    admin = db.session.get(User, admin_id)
    admin.password = generate_password_hash(new_password)
    db.session.commit()
    flash('Password changed successfully!', 'success')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/change_admin_branch/<int:admin_id>', methods=['POST'])
@login_required
def change_admin_branch(admin_id):
    if current_user.role != 'admin' or getattr(current_user,'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))
    new_branch = request.form['new_branch']
    admin = db.session.get(User, admin_id)
    admin.branch = new_branch
    db.session.commit()
    flash('Branch updated successfully!', 'success')
    return redirect(url_for('superadmin_dashboard'))

@app.route('/set_feedback_window', methods=['POST'])
@login_required
def set_feedback_window():
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    if role != 'admin':
        return redirect(url_for('home'))

    start_time = request.form.get('start_time')
    end_time = request.form.get('end_time')

    if not start_time or not end_time:
        flash('Please select both start and end date/time.', 'danger')
        return redirect(request.referrer or url_for('admin_dashboard'))

    settings = get_feedback_settings()
    settings.feedback_start = datetime.strptime(start_time, '%Y-%m-%dT%H:%M')
    settings.feedback_end = datetime.strptime(end_time, '%Y-%m-%dT%H:%M')
    settings.allow_feedback = True

    db.session.commit()
    flash('Feedback time window saved successfully!', 'success')

    return redirect(url_for('superadmin_dashboard'))



# ========================
# EXCEL UPLOAD / DELETE
# ========================
@app.route('/upload_students', methods=['POST'])
@login_required
def upload_students():
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    if 'file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(request.referrer or url_for('superadmin_dashboard'))

    file = request.files['file']
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(request.referrer or url_for('superadmin_dashboard'))

    try:
         upload_folder = os.path.join(os.getcwd(), 'uploads')
         if not os.path.exists(upload_folder):
               os.makedirs(upload_folder)

         filename = file.filename
         file_path = os.path.join(upload_folder, filename)
         file.save(file_path)

         uploaded_file = UploadedFile(
               filename=filename,
               file_type='student',
               uploaded_by=current_user.id
         )
         db.session.add(uploaded_file)
         db.session.commit()

         imported, updated = import_students_from_excel(file_path, uploaded_file.id)
         flash(f'Students imported: {imported}, updated: {updated}', 'success')

    except Exception as e:
      db.session.rollback()
      flash(f'Error: {str(e)}', 'danger')


    return redirect(request.referrer or url_for('superadmin_dashboard'))

@app.route('/upload_teachers', methods=['POST'])
@login_required
def upload_teachers():
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    if 'file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(request.referrer or url_for('admin_dashboard'))

    file = request.files['file']
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(request.referrer or url_for('admin_dashboard'))

    try:
        upload_folder = os.path.join(os.getcwd(), 'uploads')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)

        filename = file.filename
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)

        uploaded_file = UploadedFile(
            filename=filename,
            file_type='teacher',
            uploaded_by=current_user.id
        )
        db.session.add(uploaded_file)
        db.session.commit()

        df = pd.read_excel(file_path, dtype=str).fillna('')

        def normalize_col(col):
            return (
                str(col)
                .strip()
                .lower()
                .replace(" ", "_")
                .replace("-", "_")
            )

        df.columns = [normalize_col(col) for col in df.columns]

        def cell(row, *names, default=''):
            for name in names:
                key = normalize_col(name)
                if key in row:
                    value = str(row.get(key, '')).strip()
                    if value.lower() != 'nan':
                        return value
            return default

        imported_users = 0
        imported_subjects = 0
        skipped_subjects = 0

        for _, row in df.iterrows():
            username = cell(row, 'username', 'user name', 'teacher username')
            if not username:
                continue

            password = cell(row, 'password', default='teacher123') or 'teacher123'
            email = cell(row, 'email', 'e-mail', 'mail')
            name = cell(row, 'name', 'teacher name', default=username) or username
            subject = cell(row, 'subject', 'subject name')
            subject_code = cell(row, 'subject_code', 'subject code', 'code')
            branch = normalize_branch(cell(row, 'branch'))
            semester_value = cell(row, 'semester', 'sem', default='1')

            try:
                semester = int(float(semester_value))
            except:
                semester = 1

            user = User.query.filter_by(username=username).first()

            if not user:
                user = User(
                    username=username,
                    password=generate_password_hash(password),
                    role='teacher',
                    name=name,
                    email=email or f'{username.replace(" ", "").lower()}@college.com',
                    subject=subject,
                    imported_file_id=uploaded_file.id
                )
                db.session.add(user)
                db.session.flush()
                imported_users += 1
            else:
                if not user.name:
                    user.name = name
                if not user.email and email:
                    user.email = email
                if not user.subject and subject:
                    user.subject = subject

            existing_subject = TeacherSubject.query.filter_by(
                teacher_id=user.id,
                subject=subject,
                branch=branch,
                semester=semester
            ).first()

            if existing_subject:
                existing_subject.subject_code = subject_code
                skipped_subjects += 1
            else:
                teacher_subject = TeacherSubject(
                    teacher_id=user.id,
                    subject=subject,
                    subject_code=subject_code,
                    branch=branch,
                    semester=semester
                )
                db.session.add(teacher_subject)
                imported_subjects += 1

        db.session.commit()

        flash(
            f'Teachers imported: {imported_users}, subjects added: {imported_subjects}, subjects updated/skipped: {skipped_subjects}',
            'success'
        )

    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')

    return redirect(request.referrer or url_for('superadmin_dashboard'))



@app.route('/upload_admins', methods=['POST'])
@login_required
def upload_admins():
    if current_user.role != 'admin' or getattr(current_user, 'sub_role', None) != 'superadmin':
        return redirect(url_for('home'))

    if 'file' not in request.files:
        flash('No file selected!', 'danger')
        return redirect(url_for('superadmin_dashboard'))

    file = request.files['file']
    if file.filename == '':
        flash('No file selected!', 'danger')
        return redirect(url_for('superadmin_dashboard'))

    try:
        upload_folder = os.path.join(os.getcwd(), 'uploads')
        if not os.path.exists(upload_folder):
            os.makedirs(upload_folder)

        filename = file.filename
        file_path = os.path.join(upload_folder, filename)
        file.save(file_path)

        uploaded_file = UploadedFile(
            filename=filename,
            file_type='admin',
            uploaded_by=current_user.id
        )
        db.session.add(uploaded_file)
        db.session.commit()

        wb = load_workbook(file_path)
        ws = wb.active

        imported = 0
        skipped = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            username = str(row[0])

            if User.query.filter_by(username=username).first():
                skipped += 1
                continue

            user = User(
                username=username,
                password=generate_password_hash(str(row[1]) if row[1] else 'admin123'),
                role='admin',
                sub_role=str(row[2]) if len(row) > 2 and row[2] else 'branchadmin',
                branch=str(row[3]) if len(row) > 3 and row[3] else None,
                name=str(row[4]) if len(row) > 4 and row[4] else '',
                email=str(row[5]) if len(row) > 5 and row[5] else '',
                imported_file_id=uploaded_file.id
            )

            db.session.add(user)
            imported += 1

        db.session.commit()
        flash(f'Successfully imported {imported} admins! ({skipped} skipped)', 'success')

    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')

    return redirect(url_for('superadmin_dashboard'))

@app.route('/delete_file/<int:file_id>')
@login_required
def delete_file(file_id):
    role = (getattr(current_user, 'role', None) or '').strip().lower()
    if role != 'admin':
        return redirect(url_for('home'))

    file = db.session.get(UploadedFile, file_id)

    if not file:
        flash("File not found!", "danger")
        return redirect(url_for('superadmin_dashboard', tab='view'))

    try:
        users = User.query.filter_by(imported_file_id=file.id).all()
        upload_folder = os.path.join(os.getcwd(), 'uploads')
        file_path = os.path.join(upload_folder, file.filename)

        if file.file_type == 'student':
            semesters_to_delete = student_file_semesters(file_path)

            for user in users:
                feedback_query = Feedback.query.filter_by(student_id=user.id)
                if semesters_to_delete:
                    feedback_query = feedback_query.filter(Feedback.semester.in_(semesters_to_delete))
                else:
                    feedback_query = feedback_query.filter(Feedback.semester == user.semester)
                feedback_query.delete(synchronize_session=False)

                remaining_feedbacks = Feedback.query.filter_by(student_id=user.id).first()
                if remaining_feedbacks:
                    user.imported_file_id = None
                else:
                    db.session.delete(user)

        elif file.file_type == 'teacher':
            for user in users:
                Feedback.query.filter_by(teacher_id=user.id).delete(synchronize_session=False)
                TeacherSubject.query.filter_by(teacher_id=user.id).delete(synchronize_session=False)
                db.session.delete(user)

        else:
            for user in users:
                db.session.delete(user)

        if os.path.exists(file_path):
            os.remove(file_path)

        db.session.delete(file)
        db.session.commit()

        if file.file_type == 'student':
            flash("Student file and matching-semester feedbacks permanently deleted. Old/previous semester feedback is saved.", "success")
        elif file.file_type == 'teacher':
            flash("Teacher file, related teachers, subjects, and all feedback for those teachers permanently deleted!", "success")
        else:
            flash("File, related users, subjects, and feedbacks permanently deleted!", "success")

    except Exception as e:
        db.session.rollback()
        flash(f"Error deleting file: {str(e)}", "danger")

    return redirect(url_for('superadmin_dashboard', tab='view'))

# ========================
# VIEW / EDIT / DELETE STUDENTS & TEACHERS
# ========================
@app.route('/view_students')
@login_required
def view_students():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    students = User.query.filter_by(role='student', branch=current_user.branch).all() \
               if getattr(current_user, 'sub_role', None)=='branchadmin' else User.query.filter_by(role='student').all()
    return render_template('admin/view_students.html', users=students)

@app.route('/view_teachers')
@login_required
def view_teachers():
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    teachers = User.query.filter_by(role='teacher').all()
    return render_template('admin/view_teachers.html', users=teachers)

@app.route('/edit_user/<int:user_id>', methods=['POST'])
@login_required
def edit_user(user_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))
    user = db.session.get(User, user_id)
    if not user:
        flash("User not found!", "danger")
        return redirect(request.referrer)
    
    # Update fields
    user.name = request.form.get('name', user.name)
    user.email = request.form.get('email', user.email)
    if user.role=='student':
        user.branch = request.form.get('branch', user.branch)
        user.semester = int(request.form.get('semester', user.semester))
        user.academic_year = request.form.get('academic_year', user.academic_year)
        user.roll_number = request.form.get('roll_number', user.roll_number)
    elif user.role=='teacher':
        user.subject = request.form.get('subject', user.subject)
    
    db.session.commit()
    flash(f"{user.role.title()} '{user.username}' updated successfully!", "success")
    return redirect(request.referrer)


@app.route('/export_teacher_feedback')
@login_required
def export_teacher_feedback():
    if current_user.role != 'teacher':
        return redirect(url_for('home'))

    branch = request.args.get('branch')
    semester = request.args.get('semester')
    subject = request.args.get('subject')

    query = Feedback.query.join(User, Feedback.student_id == User.id).filter(
        Feedback.teacher_id == current_user.id,
        Feedback.semester == User.semester
    )

    if branch:
        query = query.filter(User.branch == branch)
    if semester:
        query = query.filter(Feedback.semester == int(semester))
    if subject:
        query = query.filter(Feedback.subject == subject)

    return export_feedbacks_response(
        query.order_by(Feedback.created_at.desc()).all(),
        f'teacher_current_feedback_{current_user.id}.xlsx'
    )
@app.route('/delete_user/<int:user_id>')
@login_required
def delete_user(user_id):
    user = db.session.get(User, user_id)

    if not user:
        flash("User not found!", "danger")
        return redirect(url_for('admin_dashboard'))

    db.session.delete(user)
    db.session.commit()

    flash("User deleted permanently!", "success")
    return redirect(url_for('admin_dashboard'))
import pandas as pd  # 👈 ye import add karna hai file ke top me


def feedbacks_for_student_file(file_path):
    df = pd.read_excel(file_path, dtype=str).fillna('')
    df.columns = [normalize_col(col) for col in df.columns]

    username_col = None
    semester_col = None

    for col in df.columns:
        normalized = normalize_col(col)
        if normalized in ['username', 'user_name', 'roll_number', 'roll_no', 'roll']:
            username_col = col
        if normalized in ['semester', 'sem']:
            semester_col = col

    if not username_col:
        return []

    usernames = [str(value).strip() for value in df[username_col].tolist() if str(value).strip()]
    students = User.query.filter(User.username.in_(usernames), User.role == 'student').all()
    student_ids = [student.id for student in students]

    if not student_ids:
        return []

    query = Feedback.query.filter(Feedback.student_id.in_(student_ids))

    if semester_col:
        semesters = sorted({parse_semester(value, None) for value in df[semester_col].tolist()})
        semesters = [sem for sem in semesters if sem is not None]
        if semesters:
            query = query.filter(Feedback.semester.in_(semesters))

    return query.order_by(Feedback.created_at.desc()).all()


@app.route('/download_file_feedback/<int:file_id>')
@login_required
def download_file_feedback(file_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    file = db.session.get(UploadedFile, file_id)
    if not file or file.file_type != 'student':
        flash("Student file not found!", "danger")
        return redirect(url_for('superadmin_dashboard', tab='view'))

    file_path = os.path.join(app.config['UPLOAD_FOLDER'], file.filename)
    if not os.path.exists(file_path):
        flash("Physical file not found!", "danger")
        return redirect(url_for('view_file', file_id=file_id))

    feedbacks = feedbacks_for_student_file(file_path)
    safe_name = os.path.splitext(file.filename)[0].replace(' ', '_')
    return export_feedbacks_response(
        feedbacks,
        f'feedback_data_{safe_name}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )
@app.route('/view_file/<int:file_id>', methods=['GET', 'POST'])
@login_required
def view_file(file_id):
    if current_user.role != 'admin':
        return redirect(url_for('home'))

    file = db.session.get(UploadedFile, file_id)
    if not file:
        flash("File not found!", "danger")
        return redirect(url_for('superadmin_dashboard'))

    file_path = upload_path(file.filename)

    if not os.path.exists(file_path):
        flash(f"Physical file missing for '{file.filename}'. Please re-upload this Excel file or delete this stale record.", "danger")
        return redirect(url_for('superadmin_dashboard', tab='view'))

    try:
        df = pd.read_excel(file_path, dtype=str).fillna('')

        if request.method == 'POST':
            action = request.form.get('action')
            columns = list(df.columns)

            if action == 'add_row':
                empty_row = {col: '' for col in columns}
                df = pd.concat([pd.DataFrame([empty_row]), df], ignore_index=True)

            elif action == 'delete_row':
                row_index = request.form.get('row_index', type=int)
                if row_index is not None and 0 <= row_index < len(df):
                    df = df.drop(df.index[row_index]).reset_index(drop=True)
                else:
                    flash("Invalid row selected!", "danger")
                    return redirect(url_for('view_file', file_id=file_id))

            elif action == 'add_col':
                new_col = request.form.get('new_column_name', '').strip()
                if not new_col:
                    flash("Column name cannot be empty!", "danger")
                    return redirect(url_for('view_file', file_id=file_id))
                if new_col in df.columns:
                    flash("Column already exists!", "danger")
                    return redirect(url_for('view_file', file_id=file_id))
                df[new_col] = ''

            elif action == 'delete_col':
                col_name = request.form.get('col_name', '').strip()
                if col_name in df.columns:
                    df = df.drop(columns=[col_name])
                else:
                    flash("Invalid column selected!", "danger")
                    return redirect(url_for('view_file', file_id=file_id))

            elif action == 'promote_students':
                if file.file_type != 'student':
                    flash("Promote is only available for student files.", "danger")
                    return redirect(url_for('view_file', file_id=file_id))

                semester_col = None
                for col in df.columns:
                    if str(col).strip().lower().replace(" ", "_") in ['semester', 'sem']:
                        semester_col = col
                        break

                if not semester_col:
                    flash("Semester column not found!", "danger")
                    return redirect(url_for('view_file', file_id=file_id))
                def promote_sem(value):
                    text = str(value).strip().lower()
                    digits = ''.join(ch for ch in text if ch.isdigit())

                    if not digits:
                        return value

                    sem = int(digits)
                    if sem >= 8:
                        return value

                    new_sem = sem + 1

                    if 'st' in text or 'nd' in text or 'rd' in text or 'th' in text:
                        return f'{new_sem}th'

                    return str(new_sem)

                df[semester_col] = df[semester_col].apply(promote_sem)

            elif action == 'save':
                row_count = int(request.form.get('row_count', 0))
                form_columns = request.form.getlist('columns')
                updated_rows = []

                for i in range(row_count):
                    row_data = {}
                    for col in form_columns:
                        row_data[col] = request.form.get(f'cell_{i}_{col}', '')
                    updated_rows.append(row_data)

                df = pd.DataFrame(updated_rows, columns=form_columns)

            df.to_excel(file_path, index=False)
            if file.file_type == 'student':
                 import_students_from_excel(file_path, file.id)
            elif file.file_type == 'teacher':
                 import_teachers_from_excel(file_path, file.id)


  


                  

          #  elif file.file_type == 'teacher':
            #     import_teachers_from_excel(file_path, file.id)
            flash("Excel file updated successfully!", "success")
            return redirect(url_for('view_file', file_id=file_id))

        return render_template('admin/view_file.html', file=file, df=df)

    except Exception as e:
        flash(f"Error opening file: {str(e)}", "danger")
        return redirect(url_for('superadmin_dashboard'))


# ========================
# INITIALIZE DATABASE
# ========================
def create_default_admin():
    with app.app_context():
        db.create_all()
        if not User.query.filter_by(username='superadmin').first():
            superadmin = User(username='superadmin', password=generate_password_hash('admin123'),
                              email='superadmin@college.com', role='admin', sub_role='superadmin',
                              name='Super Admin', branch=None)
            db.session.add(superadmin)
        for branch in BRANCHES:
            uname = branch.lower().replace('&','').replace('-','').replace(' ','')+'admin'
            if not User.query.filter_by(username=uname).first():
                admin = User(username=uname, password=generate_password_hash('admin123'),
                             email=f'{uname}@college.com', role='admin', sub_role='branchadmin',
                             name=f'{branch} Admin', branch=branch)
                db.session.add(admin)
        db.session.commit()
        print("Database initialized successfully!")
        print("\nDefault Login Credentials:")
        print("-"*50)
        print("| Role          | Username          | Password |")
        print("-"*50)
        print("| Super Admin   | superadmin        | admin123 |")
        for branch in BRANCHES:
            uname = branch.lower().replace('&','').replace('-','').replace(' ','')+'admin'
            print(f"| {branch} Admin | {uname:<15} | admin123 |")
        print("-"*50)

if __name__ == '__main__':
    create_default_admin()
    app.run(debug=True, use_reloader=False)


























